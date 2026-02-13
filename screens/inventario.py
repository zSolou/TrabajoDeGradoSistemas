import os
from PySide6 import QtCore, QtWidgets, QtGui
from core import repo, theme
from datetime import datetime

FACTORES_CONVERSION = {
    "Tablas": 30, "Tablones": 20, "Paletas": 10, "Machihembrado": 5
}

class EditarProductoDialog(QtWidgets.QDialog):
    def __init__(self, data, parent=None):
        super().__init__(parent)
        self.data = data
        self.recover_status = None # Control de recuperación
        self.setWindowTitle(f"Visualizar Lote {data.get('nro_lote', '')}")
        self.setModal(True); self.resize(500, 650)
        self.setStyleSheet(f"background-color: {theme.BG_SIDEBAR}; color: {theme.TEXT_PRIMARY};")
        self._build_ui()

    def _build_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        
        # --- MENSAJE DE SEGURIDAD ---
        info_lbl = QtWidgets.QLabel("🔒 MODO LECTURA - SOLO OBSERVACIONES EDITABLES")
        info_lbl.setStyleSheet("color: #ffa500; font-weight: bold; padding: 5px; border-bottom: 1px solid #ffa500;")
        info_lbl.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(info_lbl)

        scroll = QtWidgets.QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setStyleSheet("background-color: transparent; border: none;")
        content = QtWidgets.QWidget(); form = QtWidgets.QFormLayout(content); form.setSpacing(12)
        
        # --- CAMPOS BLOQUEADOS (READ ONLY) ---
        self.inp_lote = QtWidgets.QLineEdit(self.data.get('nro_lote', ''))
        self.inp_lote.setReadOnly(True)
        self.inp_lote.setStyleSheet(f"background-color: {theme.BG_INPUT}; border: 1px solid {theme.BORDER_COLOR}; color: #888; padding: 4px;")
        
        self.inp_qty = self._spinbox(float(self.data.get('quantity', 0)), locked=True)
        self.inp_l = self._spinbox(float(self.data.get('largo', 0)), locked=True)
        self.inp_a = self._spinbox(float(self.data.get('ancho', 0)), locked=True)
        self.inp_e = self._spinbox(float(self.data.get('espesor', 0)), locked=True)
        
        self.inp_date = QtWidgets.QDateEdit(calendarPopup=True)
        try:
            d = str(self.data.get('prod_date')).split("T")[0]
            self.inp_date.setDate(QtCore.QDate.fromString(d, "yyyy-MM-dd"))
        except: self.inp_date.setDate(QtCore.QDate.currentDate())
        self.inp_date.setReadOnly(True); self.inp_date.setDisabled(True)
        self.inp_date.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: #888; padding: 4px;")

        self.inp_calidad = QtWidgets.QLineEdit(self.data.get('quality', ''))
        self.inp_calidad.setReadOnly(True)
        self.inp_calidad.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: #888; padding: 4px;")

        # --- OBSERVACIONES (EDITABLE) ---
        self.inp_obs = QtWidgets.QPlainTextEdit(str(self.data.get('obs') or "")); self.inp_obs.setFixedHeight(80)
        self.inp_obs.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: white; border: 1px solid {theme.ACCENT_COLOR};")

        form.addRow("Nro. Lote:", self.inp_lote); form.addRow("Cantidad (Piezas):", self.inp_qty)
        form.addRow("Largo (m):", self.inp_l); form.addRow("Ancho (cm):", self.inp_a); form.addRow("Espesor (cm):", self.inp_e)
        form.addRow("Fecha Prod:", self.inp_date); form.addRow("Calidad:", self.inp_calidad)
        form.addRow("Obs (Editable):", self.inp_obs)

        scroll.setWidget(content); layout.addWidget(scroll)

        # --- SECCIÓN RECUPERAR BAJA ---
        if self.data.get('status') == 'BAJA':
            rec_frame = QtWidgets.QFrame()
            rec_frame.setStyleSheet("background-color: #3a1c1c; border-radius: 6px; padding: 10px; margin-top: 10px;")
            rl = QtWidgets.QVBoxLayout(rec_frame)
            
            lbl_baja = QtWidgets.QLabel("⚠️ ESTE PRODUCTO ESTÁ DADO DE BAJA")
            lbl_baja.setStyleSheet("color: #ff6b6b; font-weight: bold; font-size: 10pt;")
            lbl_baja.setAlignment(QtCore.Qt.AlignCenter)
            
            self.btn_rec = QtWidgets.QPushButton("♻️ RECUPERAR (Reactivar Lote)")
            self.btn_rec.setCursor(QtCore.Qt.PointingHandCursor)
            self.btn_rec.setStyleSheet(f"background-color: {theme.BTN_SUCCESS}; color: black; font-weight: bold; padding: 8px;")
            self.btn_rec.clicked.connect(self._activar_recuperacion)
            
            rl.addWidget(lbl_baja); rl.addWidget(self.btn_rec)
            layout.addWidget(rec_frame)

        # Botones Acción
        btn_box = QtWidgets.QHBoxLayout()
        btn_save = QtWidgets.QPushButton("Guardar Cambios"); btn_save.clicked.connect(self.accept)
        btn_save.setStyleSheet(f"background-color: {theme.BTN_PRIMARY}; color: white; font-weight: bold; padding: 8px;")
        
        btn_cancel = QtWidgets.QPushButton("Cancelar"); btn_cancel.clicked.connect(self.reject)
        btn_cancel.setStyleSheet(f"background-color: {theme.BTN_DANGER}; color: white; padding: 8px;")
        
        btn_box.addWidget(btn_cancel); btn_box.addWidget(btn_save); layout.addLayout(btn_box)

    def _spinbox(self, val, locked=False):
        sb = QtWidgets.QDoubleSpinBox(); sb.setRange(0, 999999); sb.setValue(val)
        if locked:
            sb.setReadOnly(True); sb.setButtonSymbols(QtWidgets.QAbstractSpinBox.NoButtons)
            sb.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: #888; border: 1px solid {theme.BORDER_COLOR}; padding: 4px;")
        else:
            sb.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: white; padding: 4px;")
        return sb

    def _activar_recuperacion(self):
        self.recover_status = "AGOTADO" # Vuelve como activo pero con stock 0
        self.btn_rec.setText("✅ SE RECUPERARÁ AL GUARDAR")
        self.btn_rec.setEnabled(False)
        self.inp_obs.appendPlainText(" [RECUPERADO DE BAJA]")

    def get_data(self):
        d = {
            "id": self.data["id"], 
            "nro_lote": self.inp_lote.text(), 
            "quantity": self.inp_qty.value(),
            "largo": self.inp_l.value(), 
            "ancho": self.inp_a.value(), 
            "espesor": self.inp_e.value(),
            "prod_date": self.inp_date.date().toString("yyyy-MM-dd"), 
            "quality": self.inp_calidad.text(), 
            "obs": self.inp_obs.toPlainText()
        }
        if self.recover_status: d["status"] = self.recover_status
        return d

class InventarioScreen(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data_existencias = []
        self.data_historial = []
        self._setup_ui()
        self.refresh()

    def _setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        
        lbl = QtWidgets.QLabel("CONTROL DE INVENTARIO Y SALIDAS")
        lbl.setStyleSheet(f"font-size: 18pt; font-weight: bold; color: {theme.ACCENT_COLOR};")
        lbl.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(lbl)

        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setStyleSheet(f"""
            QTabWidget::pane {{ border: 1px solid {theme.BORDER_COLOR}; }}
            QTabBar::tab {{ background: {theme.BG_SIDEBAR}; color: {theme.TEXT_SECONDARY}; padding: 10px 20px; font-weight: bold; }}
            QTabBar::tab:selected {{ background: {theme.BTN_PRIMARY}; color: white; }}
        """)
        
        self.tab_existencias = QtWidgets.QWidget()
        self._setup_tab_existencias(self.tab_existencias)
        self.tabs.addTab(self.tab_existencias, "📦 Existencias en Patio")
        
        self.tab_historial = QtWidgets.QWidget()
        self._setup_tab_historial(self.tab_historial)
        self.tabs.addTab(self.tab_historial, "🚛 Historial de Despachos")
        
        layout.addWidget(self.tabs)

    def _estilizar_boton(self, btn, color):
        btn.setCursor(QtCore.Qt.PointingHandCursor)
        btn.setStyleSheet(f"QPushButton {{ background-color: {color}; color: white; border-radius: 4px; padding: 6px 15px; font-weight: bold; }} QPushButton:hover {{ filter: brightness(115%); }}")

    def _setup_tab_existencias(self, parent):
        layout = QtWidgets.QVBoxLayout(parent)
        
        top_bar = QtWidgets.QHBoxLayout()
        self.search_exist = QtWidgets.QLineEdit()
        self.search_exist.setPlaceholderText("🔍 Buscar Lote, SKU...")
        self.search_exist.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: white; padding: 6px; border-radius: 4px;")
        self.search_exist.textChanged.connect(self._filtrar_existencias)
        
        self.chk_show_exhausted = QtWidgets.QCheckBox("Mostrar Agotados/Bajas")
        self.chk_show_exhausted.setStyleSheet("color: white; font-weight: bold;")
        self.chk_show_exhausted.stateChanged.connect(self.refresh)
        
        btn_edit = QtWidgets.QPushButton("👁️ Ver / Editar"); btn_edit.clicked.connect(self._editar_producto)
        btn_del = QtWidgets.QPushButton("📉 Dar de Baja"); btn_del.clicked.connect(self._dar_baja_producto)
        btn_xls = QtWidgets.QPushButton("📊 Excel"); btn_xls.clicked.connect(lambda: self._exportar_excel("existencias"))
        
        self._estilizar_boton(btn_edit, theme.BTN_PRIMARY)
        self._estilizar_boton(btn_del, theme.BTN_DANGER)
        self._estilizar_boton(btn_xls, "#217346")
        
        top_bar.addWidget(self.search_exist)
        top_bar.addWidget(self.chk_show_exhausted)
        top_bar.addWidget(btn_edit)
        top_bar.addWidget(btn_del)
        top_bar.addWidget(btn_xls)
        layout.addLayout(top_bar)

        self.table_exist = QtWidgets.QTableWidget()
        cols = ["ID", "SKU", "LOTE", "Producto", "Existencia", "Bultos", "F. Prod", "Estado", "Largo", "Ancho", "Espesor", "Calidad", "Secado", "Cepillado", "Impregnado", "Obs"]
        self.table_exist.setColumnCount(len(cols))
        self.table_exist.setHorizontalHeaderLabels(cols)
        self._estilizar_tabla(self.table_exist)
        layout.addWidget(self.table_exist)

    def _setup_tab_historial(self, parent):
        layout = QtWidgets.QVBoxLayout(parent)
        top_bar = QtWidgets.QHBoxLayout()
        self.search_hist = QtWidgets.QLineEdit()
        self.search_hist.setPlaceholderText("🔍 Buscar por Cliente, Guía...")
        self.search_hist.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: white; padding: 6px; border-radius: 4px;")
        self.search_hist.textChanged.connect(self._filtrar_historial)
        
        btn_refresh = QtWidgets.QPushButton("🔄 Actualizar"); btn_refresh.clicked.connect(self.refresh)
        btn_xls = QtWidgets.QPushButton("📊 Excel Historial"); btn_xls.clicked.connect(lambda: self._exportar_excel("historial"))
        
        self._estilizar_boton(btn_refresh, theme.BTN_PRIMARY)
        self._estilizar_boton(btn_xls, "#217346")

        top_bar.addWidget(self.search_hist)
        top_bar.addWidget(btn_refresh)
        top_bar.addWidget(btn_xls)
        layout.addLayout(top_bar)

        self.table_hist = QtWidgets.QTableWidget()
        cols = ["ID", "Fecha", "Guía", "Cliente", "Producto", "Lote", "SKU", "Cant. Salida", "Bultos Salida", "Obs"]
        self.table_hist.setColumnCount(len(cols))
        self.table_hist.setHorizontalHeaderLabels(cols)
        self._estilizar_tabla(self.table_hist)
        layout.addWidget(self.table_hist)

    def _estilizar_tabla(self, table):
        table.setStyleSheet(f"""
            QTableWidget {{ background-color: {theme.BG_SIDEBAR}; color: {theme.TEXT_PRIMARY}; gridline-color: {theme.BORDER_COLOR}; border: 1px solid {theme.BORDER_COLOR}; }}
            QHeaderView::section {{ background-color: #1b1b26; color: {theme.TEXT_SECONDARY}; padding: 5px; font-weight: bold; border: none; }}
            QTableWidget::item:selected {{ background-color: {theme.ACCENT_COLOR}; color: black; }}
        """)
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        table.horizontalHeader().setStretchLastSection(True)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

    def refresh(self):
        try:
            mostrar_todo = self.chk_show_exhausted.isChecked()
            self.data_existencias = repo.list_inventory_rows(mostrar_agotados=mostrar_todo)
            self._llenar_existencias(self.data_existencias)
            self.data_historial = repo.list_dispatches_history()
            self._llenar_historial(self.data_historial)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Error cargando datos: {e}")

    def _llenar_existencias(self, data):
        self.table_exist.setRowCount(0)
        for r in data:
            row = self.table_exist.rowCount(); self.table_exist.insertRow(row)
            
            tipo = str(r.get("product_type", "")); qty = float(r.get("quantity", 0))
            factor = FACTORES_CONVERSION.get(tipo, 1)
            bultos = int(qty / factor) if factor else 0
            
            status = r.get("status")
            if qty == 0 and status != "BAJA": status = "AGOTADO"

            vals = [str(r.get("id")), r.get("sku"), r.get("nro_lote"), tipo, f"{qty:.0f}", f"{bultos}", str(r.get("prod_date")), status, str(r.get("largo")), str(r.get("ancho")), str(r.get("espesor")), r.get("quality"), r.get("drying"), r.get("planing"), r.get("impregnated"), r.get("obs")]
            for i, v in enumerate(vals):
                it = QtWidgets.QTableWidgetItem(str(v or ""))
                if i==0: it.setData(QtCore.Qt.UserRole, r)
                if status == "BAJA": it.setForeground(QtGui.QColor("#ff6b6b"))
                elif qty == 0: it.setForeground(QtGui.QColor("gray"))
                self.table_exist.setItem(row, i, it)

    def _llenar_historial(self, data):
        self.table_hist.setRowCount(0)
        for r in data:
            row = self.table_hist.rowCount(); self.table_hist.insertRow(row)
            tipo = str(r.get("type", "")); qty = float(r.get("quantity", 0))
            factor = FACTORES_CONVERSION.get(tipo, 1); bultos = int(qty / factor) if factor else 0
            vals = [str(r.get("id")), str(r.get("date")), r.get("guide"), r.get("client"), r.get("product"), r.get("lote"), r.get("sku"), f"{qty:.0f}", f"{bultos}", r.get("obs")]
            for i, v in enumerate(vals):
                self.table_hist.setItem(row, i, QtWidgets.QTableWidgetItem(str(v or "")))

    def _filtrar_existencias(self, text):
        t = text.lower()
        res = [x for x in self.data_existencias if t in str(x.get("sku")).lower() or t in str(x.get("nro_lote")).lower() or t in str(x.get("product_type")).lower()]
        self._llenar_existencias(res)

    def _filtrar_historial(self, text):
        t = text.lower()
        res = [x for x in self.data_historial if t in str(x.get("client")).lower() or t in str(x.get("guide")).lower() or t in str(x.get("lote")).lower()]
        self._llenar_historial(res)

    def _get_selected_existencia(self):
        row = self.table_exist.currentRow()
        if row < 0: return None
        return self.table_exist.item(row, 0).data(QtCore.Qt.UserRole)

    # --- CAMBIO: AÑADIDA JUSTIFICACIÓN DE BAJA ---
    def _dar_baja_producto(self):
        data = self._get_selected_existencia()
        if not data: return
        
        if data.get("status") == "BAJA":
             QtWidgets.QMessageBox.warning(self, "Aviso", "Este producto ya está dado de BAJA.")
             return
        
        if float(data['quantity']) == 0:
            QtWidgets.QMessageBox.information(self, "Info", "Este producto ya está agotado.")
            return

        # Input Dialog para la razón
        reason, ok = QtWidgets.QInputDialog.getText(
            self, "Justificación de Baja", 
            f"Está a punto de dar de baja el Lote {data['nro_lote']}.\n\n"
            "Por favor, ingrese el motivo (Obligatorio):",
            QtWidgets.QLineEdit.Normal
        )

        if ok and reason.strip():
            repo.delete_inventory(data['id'], reason.strip())
            self.refresh()
            QtWidgets.QMessageBox.information(self, "Listo", "Producto dado de baja correctamente.")
        elif ok:
            QtWidgets.QMessageBox.warning(self, "Cancelado", "Debe ingresar un motivo para dar de baja.")

    def _editar_producto(self):
        data = self._get_selected_existencia()
        if not data: return
        dlg = EditarProductoDialog(data, self)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            repo.update_inventory(dlg.get_data())
            self.refresh()

    def _exportar_excel(self, tipo):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
        except ImportError: QtWidgets.QMessageBox.warning(self, "Error", "Instale openpyxl"); return

        filename = "inventario.xlsx" if tipo == "existencias" else "historial.xlsx"
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Guardar", filename, "Excel (*.xlsx)")
        if not path: return

        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Reporte"
            header_fill = PatternFill(start_color="1b1b26", end_color="1b1b26", fill_type="solid")
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            row_font = Font(name="Arial", size=10)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            if os.path.exists("logo.png"):
                try: img = XLImage("logo.png"); img.height = 50; img.width = 50; ws.add_image(img, "A1")
                except: pass

            ws.merge_cells("B2:E2")
            titulo = "EXISTENCIAS EN PATIO" if tipo == "existencias" else "HISTORIAL DE DESPACHOS"
            ws["B2"] = titulo; ws["B2"].font = Font(size=14, bold=True)
            ws["B3"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

            start_row = 5; table_widget = self.table_exist if tipo == "existencias" else self.table_hist
            col_count = table_widget.columnCount()

            for c in range(col_count):
                cell = ws.cell(row=start_row, column=c+1, value=table_widget.horizontalHeaderItem(c).text())
                cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin_border
                ws.column_dimensions[get_column_letter(c+1)].width = 18

            for r in range(table_widget.rowCount()):
                for c in range(col_count):
                    it = table_widget.item(r, c); txt = it.text() if it else ""
                    try: val = float(txt)
                    except: val = txt
                    cell = ws.cell(row=start_row + 1 + r, column=c+1, value=val)
                    cell.font = row_font; cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left" if isinstance(val, str) else "right")

            wb.save(path); QtWidgets.QMessageBox.information(self, "Éxito", f"Reporte guardado:\n{path}")
        except Exception as e: QtWidgets.QMessageBox.critical(self, "Error", str(e))