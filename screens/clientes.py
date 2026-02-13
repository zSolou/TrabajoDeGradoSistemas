from PySide6 import QtCore, QtWidgets, QtGui
from core import repo, theme
import re
import os
from datetime import datetime

class ClienteDialog(QtWidgets.QDialog):
    """Diálogo para Crear o Editar Cliente con Validaciones Venezolanas"""
    def __init__(self, data=None, parent=None):
        super().__init__(parent)
        self.data = data or {}
        self.setWindowTitle("Nuevo Cliente" if not data else "Editar Cliente")
        self.setModal(True)
        self.resize(450, 400)
        self.setStyleSheet(f"background-color: {theme.BG_SIDEBAR}; color: {theme.TEXT_PRIMARY};")
        self._build_ui()
        self._load_initial_data()

    def _build_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        form_layout = QtWidgets.QFormLayout()
        form_layout.setSpacing(15)

        # 1. Nombre
        self.inp_nombre = self._make_input()
        self.inp_nombre.setPlaceholderText("Ej: Inversiones El Roble C.A.")

        # 2. Documento (Selector V/J + Input Numérico)
        doc_container = QtWidgets.QWidget()
        doc_layout = QtWidgets.QHBoxLayout(doc_container)
        doc_layout.setContentsMargins(0,0,0,0)
        doc_layout.setSpacing(5)

        self.cb_doc_type = QtWidgets.QComboBox()
        self.cb_doc_type.addItems(["V", "J", "E", "G"]) 
        self.cb_doc_type.setFixedWidth(60)
        self.cb_doc_type.setStyleSheet(f"""
            QComboBox {{ background-color: {theme.BG_INPUT}; border: 1px solid {theme.BORDER_COLOR}; padding: 5px; border-radius: 4px; color: white; }}
            QComboBox::drop-down {{ border: none; }}
        """)
        self.cb_doc_type.currentTextChanged.connect(self._update_doc_constraints)

        self.inp_cedula = self._make_input()
        self.inp_cedula.setPlaceholderText("12345678")
        self.inp_cedula.setValidator(QtGui.QRegularExpressionValidator(QtCore.QRegularExpression("[0-9]*")))

        doc_layout.addWidget(self.cb_doc_type)
        doc_layout.addWidget(self.inp_cedula)

        # 3. Teléfono
        self.inp_telefono = self._make_input()
        self.inp_telefono.setPlaceholderText("0414-1234567")
        self.inp_telefono.setMaxLength(12) 
        self.inp_telefono.setValidator(QtGui.QRegularExpressionValidator(QtCore.QRegularExpression("[0-9-]*")))
        self.inp_telefono.textEdited.connect(self._format_phone)

        # 4. Email (Usuario + @ + Dominio)
        email_container = QtWidgets.QWidget()
        email_layout = QtWidgets.QHBoxLayout(email_container)
        email_layout.setContentsMargins(0,0,0,0)
        email_layout.setSpacing(5)

        self.inp_email_user = self._make_input()
        self.inp_email_user.setPlaceholderText("usuario")
        
        lbl_at = QtWidgets.QLabel("@")
        lbl_at.setStyleSheet(f"color: {theme.TEXT_PRIMARY}; font-weight: bold; font-size: 11pt;")

        self.inp_email_domain = self._make_input()
        self.inp_email_domain.setPlaceholderText("dominio.com")

        email_layout.addWidget(self.inp_email_user)
        email_layout.addWidget(lbl_at)
        email_layout.addWidget(self.inp_email_domain)

        # 5. Dirección
        self.inp_direccion = QtWidgets.QPlainTextEdit()
        self.inp_direccion.setFixedHeight(60)
        self.inp_direccion.setStyleSheet(f"background-color: {theme.BG_INPUT}; border: 1px solid {theme.BORDER_COLOR}; border-radius: 4px;")

        form_layout.addRow("Nombre / Razón Social *:", self.inp_nombre)
        form_layout.addRow("Documento (C.I. / RIF) *:", doc_container)
        form_layout.addRow("Teléfono:", self.inp_telefono)
        form_layout.addRow("Email:", email_container)
        form_layout.addRow("Dirección:", self.inp_direccion)

        layout.addLayout(form_layout)

        # Botones
        btn_layout = QtWidgets.QHBoxLayout()
        btn_save = QtWidgets.QPushButton("Guardar Datos")
        btn_save.setCursor(QtCore.Qt.PointingHandCursor)
        btn_save.setStyleSheet(f"background-color: {theme.BTN_SUCCESS}; color: black; font-weight: bold; padding: 10px 20px; border-radius: 4px;")
        btn_save.clicked.connect(self._validate_and_accept)
        
        btn_cancel = QtWidgets.QPushButton("Cancelar")
        btn_cancel.setCursor(QtCore.Qt.PointingHandCursor)
        btn_cancel.setStyleSheet(f"background-color: {theme.BTN_DANGER}; color: white; padding: 10px 20px; border-radius: 4px;")
        btn_cancel.clicked.connect(self.reject)

        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

        self._update_doc_constraints(self.cb_doc_type.currentText())

    def _make_input(self):
        inp = QtWidgets.QLineEdit()
        inp.setStyleSheet(f"background-color: {theme.BG_INPUT}; border: 1px solid {theme.BORDER_COLOR}; padding: 5px; border-radius: 4px; color: white;")
        return inp

    def _load_initial_data(self):
        if not self.data: return
        
        self.inp_nombre.setText(self.data.get("name", ""))
        
        email_full = self.data.get("email", "")
        if "@" in email_full:
            parts = email_full.split("@")
            self.inp_email_user.setText(parts[0])
            if len(parts) > 1: self.inp_email_domain.setText(parts[1])
        else:
            self.inp_email_user.setText(email_full)

        self.inp_telefono.setText(self.data.get("phone", ""))
        self.inp_direccion.setPlainText(self.data.get("address", ""))

        doc_full = self.data.get("document_id", "")
        if "-" in doc_full:
            parts = doc_full.split("-")
            if len(parts) == 2 and parts[0] in ["V", "J", "E", "G"]:
                self.cb_doc_type.setCurrentText(parts[0])
                self.inp_cedula.setText(parts[1])
            else:
                self.inp_cedula.setText(doc_full)
        else:
            self.inp_cedula.setText(doc_full)

    def _update_doc_constraints(self, tipo):
        self.inp_cedula.clear()
        if tipo == "V" or tipo == "E":
            self.inp_cedula.setMaxLength(8)
            self.inp_cedula.setPlaceholderText("Máx 8 dígitos")
        else: 
            self.inp_cedula.setMaxLength(10)
            self.inp_cedula.setPlaceholderText("Máx 10 dígitos")

    def _format_phone(self, text):
        clean_text = text.replace("-", "")
        if len(text) < len(self.inp_telefono.property("last_text") or ""):
            self.inp_telefono.setProperty("last_text", text)
            return
        formatted = clean_text
        if len(clean_text) > 4: formatted = clean_text[:4] + "-" + clean_text[4:]
        if text != formatted:
            self.inp_telefono.setText(formatted)
            self.inp_telefono.setCursorPosition(len(formatted))
        self.inp_telefono.setProperty("last_text", formatted)

    def _validate_and_accept(self):
        if not self.inp_nombre.text().strip():
            QtWidgets.QMessageBox.warning(self, "Error", "El Nombre es obligatorio.")
            return
        
        cedula = self.inp_cedula.text().strip()
        if not cedula:
            QtWidgets.QMessageBox.warning(self, "Error", "El Documento es obligatorio.")
            return

        tipo_doc = self.cb_doc_type.currentText()
        if tipo_doc in ["J", "G"] and len(cedula) < 8:
            QtWidgets.QMessageBox.warning(self, "Error", f"Para RIF ({tipo_doc}), el número debe tener al menos 8 dígitos.")
            return
        
        user_part = self.inp_email_user.text().strip()
        domain_part = self.inp_email_domain.text().strip()
        
        if user_part or domain_part:
            if not user_part:
                QtWidgets.QMessageBox.warning(self, "Error", "Falta el usuario del correo.")
                self.inp_email_user.setFocus(); return
            if not domain_part:
                QtWidgets.QMessageBox.warning(self, "Error", "Falta el dominio del correo.")
                self.inp_email_domain.setFocus(); return
            if "." not in domain_part:
                QtWidgets.QMessageBox.warning(self, "Error", "El dominio debe tener un punto (ej: .com).")
                self.inp_email_domain.setFocus(); return

        self.accept()

    def get_data(self):
        full_doc = f"{self.cb_doc_type.currentText()}-{self.inp_cedula.text().strip()}"
        email_full = ""
        u = self.inp_email_user.text().strip()
        d = self.inp_email_domain.text().strip()
        if u and d: email_full = f"{u}@{d}"
        
        return {
            "nombre": self.inp_nombre.text().strip(),
            "cedula_rif": full_doc,
            "telefono": self.inp_telefono.text().strip(),
            "email": email_full,
            "direccion": self.inp_direccion.toPlainText().strip()
        }


class ClientesScreen(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()
        self.refresh()

    def _setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # --- Encabezado ---
        header = QtWidgets.QHBoxLayout()
        lbl = QtWidgets.QLabel("GESTIÓN DE CLIENTES")
        lbl.setStyleSheet(f"font-size: 18pt; font-weight: bold; color: {theme.ACCENT_COLOR};")
        header.addWidget(lbl)
        header.addStretch()
        
        self.chk_ver_inactivos = QtWidgets.QCheckBox("Ver Desactivados")
        self.chk_ver_inactivos.setStyleSheet(f"color: {theme.TEXT_SECONDARY}; font-weight: bold;")
        self.chk_ver_inactivos.stateChanged.connect(self.refresh)
        header.addWidget(self.chk_ver_inactivos)
        
        btn_nuevo = QtWidgets.QPushButton("+ Nuevo Cliente")
        btn_nuevo.setCursor(QtCore.Qt.PointingHandCursor)
        btn_nuevo.setStyleSheet(f"""
            QPushButton {{ background-color: {theme.BTN_SUCCESS}; color: black; border-radius: 5px; padding: 6px 15px; font-weight: bold; }}
            QPushButton:hover {{ background-color: #00cfa5; }}
        """)
        btn_nuevo.clicked.connect(self._nuevo_cliente)
        header.addWidget(btn_nuevo)
        layout.addLayout(header)

        # --- Barra de Acciones ---
        actions_layout = QtWidgets.QHBoxLayout()
        btn_edit = QtWidgets.QPushButton("✎ Editar Seleccionado")
        btn_edit.setCursor(QtCore.Qt.PointingHandCursor)
        btn_edit.setStyleSheet(f"background-color: {theme.BTN_PRIMARY}; color: white; border-radius: 4px; padding: 6px 12px;")
        btn_edit.clicked.connect(self._editar_cliente)
        
        self.btn_toggle = QtWidgets.QPushButton("Ban/Unban") 
        self.btn_toggle.setCursor(QtCore.Qt.PointingHandCursor)
        self.btn_toggle.clicked.connect(self._toggle_activo)
        
        # --- NUEVO BOTÓN: EXPORTAR ---
        btn_xls = QtWidgets.QPushButton("📊 Exportar Excel")
        btn_xls.setCursor(QtCore.Qt.PointingHandCursor)
        btn_xls.setStyleSheet("background-color: #217346; color: white; border-radius: 4px; padding: 6px 12px; font-weight: bold;")
        btn_xls.clicked.connect(self._exportar_excel)
        # -----------------------------

        actions_layout.addWidget(btn_edit)
        actions_layout.addWidget(self.btn_toggle)
        actions_layout.addWidget(btn_xls) # Añadido al layout
        actions_layout.addStretch()
        layout.addLayout(actions_layout)

        # --- Tabla ---
        self.table = QtWidgets.QTableWidget()
        columns = ["ID", "Nombre", "Documento", "Teléfono", "Email", "Estado"]
        self.table.setColumnCount(len(columns))
        self.table.setHorizontalHeaderLabels(columns)
        self.table.setStyleSheet(f"""
            QTableWidget {{ background-color: {theme.BG_SIDEBAR}; color: white; border: 1px solid {theme.BORDER_COLOR}; border-radius: 5px; }}
            QHeaderView::section {{ background-color: #1b1b26; color: {theme.TEXT_SECONDARY}; padding: 5px; border: none; font-weight: bold; }}
            QTableWidget::item:selected {{ background-color: {theme.ACCENT_COLOR}; color: black; }}
        """)
        self.table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.itemSelectionChanged.connect(self._update_buttons)
        layout.addWidget(self.table)
        
        self._update_buttons()

    def refresh(self):
        self.table.setRowCount(0)
        ver_todos = self.chk_ver_inactivos.isChecked()
        try:
            clientes = repo.list_clients(solo_activos=not ver_todos)
            
            for c in clientes:
                r = self.table.rowCount()
                self.table.insertRow(r)
                
                is_active = getattr(c, "is_active", True)
                estado_str = "ACTIVO" if is_active else "INACTIVO"
                
                items = [str(c.id), c.name, c.document_id, c.phone or "", c.email or "", estado_str]
                
                for i, val in enumerate(items):
                    it = QtWidgets.QTableWidgetItem(val)
                    if not is_active: it.setForeground(QtGui.QColor("#ff6b6b")) 
                    if i == 0: it.setData(QtCore.Qt.UserRole, c)
                    self.table.setItem(r, i, it)
                    
        except Exception as e: print(f"Error clientes: {e}")

    def _get_selected_client(self):
        row = self.table.currentRow()
        if row < 0: return None
        return self.table.item(row, 0).data(QtCore.Qt.UserRole)

    def _update_buttons(self):
        client = self._get_selected_client()
        if not client:
            self.btn_toggle.setEnabled(False)
            self.btn_toggle.setText("Desactivar / Activar")
            self.btn_toggle.setStyleSheet(f"background-color: gray; color: white; border-radius: 4px; padding: 6px 12px;")
            return

        self.btn_toggle.setEnabled(True)
        is_active = getattr(client, "is_active", True)
        
        if is_active:
            self.btn_toggle.setText("🚫 Desactivar Cliente")
            self.btn_toggle.setStyleSheet(f"background-color: {theme.BTN_DANGER}; color: white; border-radius: 4px; padding: 6px 12px;")
        else:
            self.btn_toggle.setText("✅ Reactivar Cliente")
            self.btn_toggle.setStyleSheet(f"background-color: {theme.BTN_SUCCESS}; color: black; border-radius: 4px; padding: 6px 12px;")

    def _nuevo_cliente(self):
        dialog = ClienteDialog(parent=self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            data = dialog.get_data()
            try:
                repo.create_client(data)
                QtWidgets.QMessageBox.information(self, "Éxito", "Cliente registrado correctamente.")
                self.refresh()
            except Exception as e: QtWidgets.QMessageBox.critical(self, "Error", str(e))

    def _editar_cliente(self):
        client = self._get_selected_client()
        if not client: return
        data_dict = {"name": client.name, "document_id": client.document_id, "phone": client.phone, "email": client.email, "address": client.address}
        dialog = ClienteDialog(data=data_dict, parent=self)
        if dialog.exec_() == QtWidgets.QDialog.Accepted:
            new_data = dialog.get_data()
            try:
                repo.update_client(client.id, new_data)
                QtWidgets.QMessageBox.information(self, "Éxito", "Cliente actualizado.")
                self.refresh()
            except Exception as e: QtWidgets.QMessageBox.critical(self, "Error", str(e))

    def _toggle_activo(self):
        client = self._get_selected_client()
        if not client: return
        is_active = getattr(client, "is_active", True)
        action = "desactivar" if is_active else "reactivar"
        confirm = QtWidgets.QMessageBox.question(self, f"Confirmar {action}", f"¿Desea {action} al cliente '{client.name}'?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if confirm == QtWidgets.QMessageBox.Yes:
            try:
                repo.toggle_client_active(client.id, not is_active)
                self.refresh(); self._update_buttons()
            except Exception as e: QtWidgets.QMessageBox.critical(self, "Error", str(e))

    # --- FUNCIÓN DE EXPORTACIÓN ---
    def _exportar_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
        except ImportError:
            QtWidgets.QMessageBox.warning(self, "Error", "Instale openpyxl para exportar.")
            return

        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Exportar Clientes", "clientes.xlsx", "Excel (*.xlsx)")
        if not path: return

        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Cartera Clientes"
            
            # Estilos
            header_fill = PatternFill(start_color="1b1b26", end_color="1b1b26", fill_type="solid")
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            row_font = Font(name="Arial", size=10)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Logo
            if os.path.exists("logo.png"):
                try: img = XLImage("logo.png"); img.height = 50; img.width = 50; ws.add_image(img, "A1")
                except: pass

            # Título
            ws.merge_cells("B2:E2")
            ws["B2"] = "CARTERA DE CLIENTES REGISTRADOS"
            ws["B2"].font = Font(size=14, bold=True)
            ws["B3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y')}"

            start_row = 5
            headers = ["ID", "Nombre / Razón Social", "Documento", "Teléfono", "Email", "Estado"]
            
            # Escribir Encabezados
            for i, h in enumerate(headers):
                c = ws.cell(row=start_row, column=i+1, value=h)
                c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center"); c.border = thin_border
                ws.column_dimensions[get_column_letter(i+1)].width = 25

            # Escribir Filas
            for r in range(self.table.rowCount()):
                for c in range(len(headers)):
                    it = self.table.item(r, c)
                    val = it.text() if it else ""
                    # Convertir ID a número
                    if c == 0: 
                        try: val = int(val)
                        except: pass
                    
                    cell = ws.cell(row=start_row + 1 + r, column=c+1, value=val)
                    cell.font = row_font; cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left")

            wb.save(path)
            QtWidgets.QMessageBox.information(self, "Éxito", f"Reporte guardado en:\n{path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Fallo al exportar: {e}")