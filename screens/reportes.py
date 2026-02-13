from PySide6 import QtCore, QtWidgets, QtGui
from datetime import date, timedelta
from core import repo, theme
import sys

# Factores
FACTORES_CONVERSION = { "Tablas": 30, "Tablones": 20, "Paletas": 10, "Machihembrado": 5 }

# --- MATPLOTLIB ---
try:
    import matplotlib
    matplotlib.use('Qt5Agg')
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=3.5, height=3.5, dpi=100):
        fig = Figure(figsize=(width, height), dpi=dpi)
        fig.patch.set_facecolor(theme.BG_SIDEBAR)
        self.axes = fig.add_subplot(111)
        super().__init__(fig)

def exportar_tabla_excel(parent, table_widget, filename_base):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        QtWidgets.QMessageBox.warning(parent, "Error", "Instale openpyxl.")
        return

    path, _ = QtWidgets.QFileDialog.getSaveFileName(parent, "Exportar", f"{filename_base}.xlsx", "Excel (*.xlsx)")
    if not path: return

    try:
        wb = openpyxl.Workbook(); ws = wb.active
        
        # Estilos Excel
        header_fill = PatternFill(start_color="1b1b26", end_color="1b1b26", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = [table_widget.horizontalHeaderItem(c).text() for c in range(table_widget.columnCount())]
        
        # Escribir Headers
        for i, h in enumerate(headers):
            cell = ws.cell(row=1, column=i+1, value=h)
            cell.fill = header_fill; cell.font = header_font; cell.border = thin_border
            ws.column_dimensions[get_column_letter(i+1)].width = 18

        # Escribir Datos
        for r in range(table_widget.rowCount()):
            for c in range(table_widget.columnCount()):
                it = table_widget.item(r, c)
                txt = it.text() if it else ""
                try: val = float(txt)
                except: val = txt
                cell = ws.cell(row=r+2, column=c+1, value=val)
                cell.border = thin_border

        wb.save(path); QtWidgets.QMessageBox.information(parent, "Éxito", f"Guardado: {path}")
    except Exception as e: QtWidgets.QMessageBox.critical(parent, "Error", str(e))

class ReportesScreen(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._setup_ui()

    def _setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        t = QtWidgets.QLabel("CENTRO DE REPORTES")
        t.setStyleSheet(f"font-size: 18pt; font-weight: bold; color: {theme.ACCENT_COLOR}; margin-bottom: 10px;")
        t.setAlignment(QtCore.Qt.AlignCenter)
        layout.addWidget(t)

        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setStyleSheet(f"""
            QTabWidget::pane {{ border: 1px solid {theme.BORDER_COLOR}; }} 
            QTabBar::tab {{ background: {theme.BG_SIDEBAR}; color: {theme.TEXT_SECONDARY}; padding: 8px 25px; font-weight: bold; margin-right: 2px; }} 
            QTabBar::tab:selected {{ background: {theme.BTN_PRIMARY}; color: white; }}
        """)

        self.tab_prod = QtWidgets.QWidget(); self._setup_prod_tab(self.tab_prod)
        self.tabs.addTab(self.tab_prod, "🏭 Producción")

        self.tab_disp = QtWidgets.QWidget(); self._setup_disp_tab(self.tab_disp)
        self.tabs.addTab(self.tab_disp, "🚚 Despachos")

        self.tab_lote = QtWidgets.QWidget(); self._setup_lote_tab(self.tab_lote)
        self.tabs.addTab(self.tab_lote, "🔢 Por Lotes")

        layout.addWidget(self.tabs)

    def _estilizar_input(self, widget):
        widget.setStyleSheet(f"background-color: {theme.BG_INPUT}; color: white; padding: 5px; border: 1px solid {theme.BORDER_COLOR}; border-radius: 4px; min-width: 100px;")

    def _set_date_range(self, d1_widget, d2_widget, mode):
        today = date.today()
        if mode == "week":
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
        elif mode == "month":
            start = today.replace(day=1)
            next_month = today.replace(day=28) + timedelta(days=4)
            end = next_month.replace(day=1) - timedelta(days=1)
        elif mode == "all":
            start = date(2000, 1, 1); end = date(2030, 12, 31)
        d1_widget.setDate(start); d2_widget.setDate(end)

    # ---------------- TAB 1: PRODUCCIÓN ----------------
    def _setup_prod_tab(self, parent):
        l = QtWidgets.QVBoxLayout(parent)
        filter_box = QtWidgets.QGroupBox("Filtros"); filter_box.setStyleSheet(f"color: white; border: 1px solid {theme.BORDER_COLOR}; padding: 10px;")
        fl = QtWidgets.QVBoxLayout(filter_box)

        # Fila 1
        r1 = QtWidgets.QHBoxLayout()
        self.d1_prod = QtWidgets.QDateEdit(date.today().replace(day=1)); self.d1_prod.setCalendarPopup(True)
        self.d2_prod = QtWidgets.QDateEdit(date.today()); self.d2_prod.setCalendarPopup(True)
        self._estilizar_input(self.d1_prod); self._estilizar_input(self.d2_prod)
        
        btn_w = QtWidgets.QPushButton("Semana"); btn_w.clicked.connect(lambda: self._set_date_range(self.d1_prod, self.d2_prod, "week"))
        btn_m = QtWidgets.QPushButton("Mes"); btn_m.clicked.connect(lambda: self._set_date_range(self.d1_prod, self.d2_prod, "month"))
        btn_a = QtWidgets.QPushButton("Todos"); btn_a.clicked.connect(lambda: self._set_date_range(self.d1_prod, self.d2_prod, "all"))
        for b in [btn_w, btn_m, btn_a]: b.setStyleSheet("background-color: #444; color: white; padding: 4px 8px; border-radius: 4px;")

        r1.addWidget(QtWidgets.QLabel("Desde:")); r1.addWidget(self.d1_prod)
        r1.addWidget(QtWidgets.QLabel("Hasta:")); r1.addWidget(self.d2_prod)
        r1.addWidget(btn_w); r1.addWidget(btn_m); r1.addWidget(btn_a); r1.addStretch()

        # Fila 2
        r2 = QtWidgets.QHBoxLayout()
        self.cb_prod_filter = QtWidgets.QComboBox(); self.cb_prod_filter.addItems(["Todos los Productos", "Tablas", "Machihembrado", "Tablones", "Paletas"])
        self.cb_qual_filter = QtWidgets.QComboBox(); self.cb_qual_filter.addItems(["Todas", "Tipo 1", "Tipo 2", "Tipo 3", "Tipo 4"])
        self._estilizar_input(self.cb_prod_filter); self._estilizar_input(self.cb_qual_filter)

        btn_s = QtWidgets.QPushButton("🔍 Buscar"); btn_s.clicked.connect(self._search_prod)
        btn_s.setStyleSheet(f"background-color: {theme.BTN_PRIMARY}; font-weight: bold; padding: 6px 20px; border-radius: 4px;")

        r2.addWidget(QtWidgets.QLabel("Prod:")); r2.addWidget(self.cb_prod_filter)
        r2.addWidget(QtWidgets.QLabel("Calidad:")); r2.addWidget(self.cb_qual_filter)
        r2.addWidget(btn_s); r2.addStretch()

        fl.addLayout(r1); fl.addLayout(r2); l.addWidget(filter_box)

        spl = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.table_prod = QtWidgets.QTableWidget()
        self.table_prod.setColumnCount(8); self.table_prod.setHorizontalHeaderLabels(["Fecha", "Lote", "SKU", "Producto", "Calidad", "Cant.", "Bultos", "Estado"])
        self._style_table(self.table_prod)
        spl.addWidget(self.table_prod)

        if MATPLOTLIB_AVAILABLE:
            self.chart_prod = MplCanvas(self, width=3.5, height=3.5, dpi=90)
            spl.addWidget(self.chart_prod)
        l.addWidget(spl)

        btn_xls = QtWidgets.QPushButton("📊 Exportar Excel"); btn_xls.clicked.connect(lambda: exportar_tabla_excel(self, self.table_prod, "produccion"))
        btn_xls.setStyleSheet("background-color: #217346; color: white; padding: 8px; font-weight: bold;"); l.addWidget(btn_xls)

    def _search_prod(self):
        d1 = self.d1_prod.date().toPython(); d2 = self.d2_prod.date().toPython()
        
        # --- VALIDACIÓN DE FECHAS ---
        if d1 > d2:
            QtWidgets.QMessageBox.warning(self, "Fecha Inválida", "La fecha 'Desde' no puede ser mayor que 'Hasta'.")
            return
        # ----------------------------

        pname = self.cb_prod_filter.currentText(); pname = "" if "Todos" in pname else pname
        qual = self.cb_qual_filter.currentText()

        try:
            data = repo.report_production_period(d1, d2, pname, qual)
            self.table_prod.setRowCount(0); stats = {}
            for r in data:
                row = self.table_prod.rowCount(); self.table_prod.insertRow(row)
                tipo = r['producto']; pzas = r['piezas_iniciales']
                factor = FACTORES_CONVERSION.get(tipo, 1)
                
                self.table_prod.setItem(row, 0, QtWidgets.QTableWidgetItem(str(r['fecha'])))
                self.table_prod.setItem(row, 1, QtWidgets.QTableWidgetItem(str(r['lote'])))
                self.table_prod.setItem(row, 2, QtWidgets.QTableWidgetItem(str(r['sku'])))
                self.table_prod.setItem(row, 3, QtWidgets.QTableWidgetItem(str(tipo)))
                self.table_prod.setItem(row, 4, QtWidgets.QTableWidgetItem(str(r.get('quality', '-'))))
                self.table_prod.setItem(row, 5, QtWidgets.QTableWidgetItem(f"{pzas:.0f}"))
                self.table_prod.setItem(row, 6, QtWidgets.QTableWidgetItem(f"{pzas/factor:.1f}" if factor else "0"))
                self.table_prod.setItem(row, 7, QtWidgets.QTableWidgetItem(str(r['status'])))
                stats[tipo] = stats.get(tipo, 0) + pzas
            if MATPLOTLIB_AVAILABLE: self._update_chart(self.chart_prod, stats, "Producción (Piezas)")
        except Exception as e: QtWidgets.QMessageBox.critical(self, "Error", str(e))

    # ---------------- TAB 2: DESPACHOS ----------------
    def _setup_disp_tab(self, parent):
        l = QtWidgets.QVBoxLayout(parent)
        filter_box = QtWidgets.QGroupBox("Filtros"); filter_box.setStyleSheet(f"color: white; border: 1px solid {theme.BORDER_COLOR}; padding: 10px;")
        fl = QtWidgets.QVBoxLayout(filter_box)

        r1 = QtWidgets.QHBoxLayout()
        self.d1_disp = QtWidgets.QDateEdit(date.today().replace(day=1)); self.d1_disp.setCalendarPopup(True)
        self.d2_disp = QtWidgets.QDateEdit(date.today()); self.d2_disp.setCalendarPopup(True)
        self._estilizar_input(self.d1_disp); self._estilizar_input(self.d2_disp)
        
        btn_w = QtWidgets.QPushButton("Semana"); btn_w.clicked.connect(lambda: self._set_date_range(self.d1_disp, self.d2_disp, "week"))
        btn_m = QtWidgets.QPushButton("Mes"); btn_m.clicked.connect(lambda: self._set_date_range(self.d1_disp, self.d2_disp, "month"))
        btn_a = QtWidgets.QPushButton("Todos"); btn_a.clicked.connect(lambda: self._set_date_range(self.d1_disp, self.d2_disp, "all"))
        for b in [btn_w, btn_m, btn_a]: b.setStyleSheet("background-color: #444; color: white; padding: 4px 8px; border-radius: 4px;")

        r1.addWidget(QtWidgets.QLabel("Desde:")); r1.addWidget(self.d1_disp)
        r1.addWidget(QtWidgets.QLabel("Hasta:")); r1.addWidget(self.d2_disp)
        r1.addWidget(btn_w); r1.addWidget(btn_m); r1.addWidget(btn_a); r1.addStretch()

        r2 = QtWidgets.QHBoxLayout()
        self.cb_client = QtWidgets.QComboBox(); self.cb_client.addItem("Todos", None)
        for c in repo.list_clients(): self.cb_client.addItem(c.name, c.id)
        self.cb_disp_prod = QtWidgets.QComboBox(); self.cb_disp_prod.addItems(["Todos los Productos", "Tablas", "Machihembrado", "Tablones", "Paletas"])
        self.txt_guide = QtWidgets.QLineEdit(); self.txt_guide.setPlaceholderText("Nro. Guía"); self._estilizar_input(self.txt_guide)
        self._estilizar_input(self.cb_client); self._estilizar_input(self.cb_disp_prod)

        btn_s = QtWidgets.QPushButton("🔍 Buscar"); btn_s.clicked.connect(self._search_disp)
        btn_s.setStyleSheet(f"background-color: {theme.BTN_PRIMARY}; font-weight: bold; padding: 6px 20px; border-radius: 4px;")

        r2.addWidget(QtWidgets.QLabel("Cliente:")); r2.addWidget(self.cb_client)
        r2.addWidget(QtWidgets.QLabel("Prod:")); r2.addWidget(self.cb_disp_prod)
        r2.addWidget(QtWidgets.QLabel("Guía:")); r2.addWidget(self.txt_guide)
        r2.addWidget(btn_s); r2.addStretch()

        fl.addLayout(r1); fl.addLayout(r2); l.addWidget(filter_box)

        spl = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        self.table_disp = QtWidgets.QTableWidget()
        self.table_disp.setColumnCount(9); self.table_disp.setHorizontalHeaderLabels(["Fecha", "Guía", "Cliente", "Producto", "Lote", "SKU", "Cant.", "Bultos", "Obs"])
        self._style_table(self.table_disp)
        spl.addWidget(self.table_disp)

        if MATPLOTLIB_AVAILABLE:
            self.chart_disp = MplCanvas(self, width=3.5, height=3.5, dpi=90)
            spl.addWidget(self.chart_disp)
        l.addWidget(spl)

        btn_xls = QtWidgets.QPushButton("📊 Exportar Excel"); btn_xls.clicked.connect(lambda: exportar_tabla_excel(self, self.table_disp, "despachos"))
        btn_xls.setStyleSheet("background-color: #217346; color: white; padding: 8px; font-weight: bold;"); l.addWidget(btn_xls)

    def _search_disp(self):
        d1 = self.d1_disp.date().toPython(); d2 = self.d2_disp.date().toPython()
        
        # --- VALIDACIÓN ---
        if d1 > d2:
            QtWidgets.QMessageBox.warning(self, "Fecha Inválida", "La fecha 'Desde' no puede ser mayor que 'Hasta'.")
            return
        # ------------------

        cid = self.cb_client.currentData()
        pname = self.cb_disp_prod.currentText(); pname = "" if "Todos" in pname else pname
        guide = self.txt_guide.text().strip()

        try:
            data = repo.report_dispatches_detailed(d1, d2, cid, pname, guide)
            self.table_disp.setRowCount(0); stats = {}
            for r in data:
                row = self.table_disp.rowCount(); self.table_disp.insertRow(row)
                tipo = str(r['producto']); pzas = r['cantidad']
                factor = FACTORES_CONVERSION.get(tipo, 1)
                bultos = pzas / factor if factor else 0

                self.table_disp.setItem(row, 0, QtWidgets.QTableWidgetItem(str(r['fecha'])))
                self.table_disp.setItem(row, 1, QtWidgets.QTableWidgetItem(str(r['guia'])))
                self.table_disp.setItem(row, 2, QtWidgets.QTableWidgetItem(str(r['cliente'])))
                self.table_disp.setItem(row, 3, QtWidgets.QTableWidgetItem(tipo))
                self.table_disp.setItem(row, 4, QtWidgets.QTableWidgetItem(str(r['lote'])))
                self.table_disp.setItem(row, 5, QtWidgets.QTableWidgetItem(str(r['sku'])))
                self.table_disp.setItem(row, 6, QtWidgets.QTableWidgetItem(f"{pzas:.0f}"))
                self.table_disp.setItem(row, 7, QtWidgets.QTableWidgetItem(f"{bultos:.1f}"))
                self.table_disp.setItem(row, 8, QtWidgets.QTableWidgetItem(str(r['obs'])))
                stats[tipo] = stats.get(tipo, 0) + pzas
            if MATPLOTLIB_AVAILABLE: self._update_chart(self.chart_disp, stats, "Despachos (Piezas)")
        except Exception as e: QtWidgets.QMessageBox.critical(self, "Error", str(e))

    # ---------------- TAB 3: LOTES ----------------
    def _setup_lote_tab(self, parent):
        l = QtWidgets.QVBoxLayout(parent)
        h = QtWidgets.QHBoxLayout(); h.setSpacing(15)
        
        self.s_l1 = QtWidgets.QSpinBox(); self.s_l1.setRange(0, 999); self.s_l1.setPrefix("Lote ")
        self.s_l2 = QtWidgets.QSpinBox(); self.s_l2.setRange(0, 999); self.s_l2.setPrefix("Lote ")
        self._estilizar_input(self.s_l1); self._estilizar_input(self.s_l2)

        self.cb_lote_prod = QtWidgets.QComboBox(); self.cb_lote_prod.addItems(["Todos los Productos", "Tablas", "Machihembrado", "Tablones", "Paletas"])
        self._estilizar_input(self.cb_lote_prod)
        self.chk_agotados = QtWidgets.QCheckBox("Incluir Agotados"); self.chk_agotados.setStyleSheet("color: white; font-weight: bold;")

        btn = QtWidgets.QPushButton("🔍 Buscar"); btn.clicked.connect(self._search_lotes)
        btn.setStyleSheet(f"background-color: {theme.BTN_PRIMARY}; font-weight: bold; padding: 6px 15px; border-radius: 4px; color: white;")
        
        h.addWidget(QtWidgets.QLabel("Desde:")); h.addWidget(self.s_l1)
        h.addWidget(QtWidgets.QLabel("Hasta:")); h.addWidget(self.s_l2)
        h.addWidget(self.cb_lote_prod); h.addWidget(self.chk_agotados); h.addWidget(btn); h.addStretch()
        l.addLayout(h)

        self.table_lote = QtWidgets.QTableWidget()
        self.table_lote.setColumnCount(7); self.table_lote.setHorizontalHeaderLabels(["Lote", "SKU", "Producto", "F. Prod", "Stock (Pzas)", "Bultos", "Estado"])
        self._style_table(self.table_lote); l.addWidget(self.table_lote)

    def _search_lotes(self):
        l1 = self.s_l1.value(); l2 = self.s_l2.value()
        
        # --- VALIDACIÓN DE RANGO ---
        if l1 > l2: 
            QtWidgets.QMessageBox.warning(self, "Error", "El lote inicial no puede ser mayor al final.")
            return
        # ---------------------------

        incluir = self.chk_agotados.isChecked()
        pname = self.cb_lote_prod.currentText(); pname = None if "Todos" in pname else pname

        try:
            data = repo.report_by_lot_range(l1, l2, incluir, pname)
            self.table_lote.setRowCount(0)
            for r in data:
                row = self.table_lote.rowCount(); self.table_lote.insertRow(row)
                tipo = str(r['producto']); stock = r['stock_actual']
                factor = FACTORES_CONVERSION.get(tipo, 1)
                bultos = stock / factor if factor else 0

                self.table_lote.setItem(row, 0, QtWidgets.QTableWidgetItem(str(r['lote'])))
                self.table_lote.setItem(row, 1, QtWidgets.QTableWidgetItem(str(r['sku'])))
                self.table_lote.setItem(row, 2, QtWidgets.QTableWidgetItem(tipo))
                self.table_lote.setItem(row, 3, QtWidgets.QTableWidgetItem(str(r['fecha_prod'])))
                self.table_lote.setItem(row, 4, QtWidgets.QTableWidgetItem(f"{stock:.0f}"))
                self.table_lote.setItem(row, 5, QtWidgets.QTableWidgetItem(f"{bultos:.1f}"))
                self.table_lote.setItem(row, 6, QtWidgets.QTableWidgetItem(str(r['estado'])))
        except Exception as e: QtWidgets.QMessageBox.critical(self, "Error", str(e))

    def _style_table(self, t):
        t.setStyleSheet(f"QTableWidget {{ background-color: {theme.BG_SIDEBAR}; color: {theme.TEXT_PRIMARY}; gridline-color: {theme.BORDER_COLOR}; }} QHeaderView::section {{ background-color: #1b1b26; color: {theme.TEXT_SECONDARY}; padding: 8px; font-weight: bold; }} QTableWidget::item {{ padding: 5px; }}")
        t.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        t.verticalHeader().setVisible(False); t.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

    def _update_chart(self, canvas, data_dict, title):
        canvas.axes.clear()
        if not data_dict: canvas.axes.text(0.5, 0.5, "Sin datos", ha='center', va='center', color='white'); canvas.draw(); return
        labels = list(data_dict.keys()); sizes = list(data_dict.values())
        colors = ['#00f2c3', '#fd5d93', '#ffc107', '#1d8cf8', '#e14eca']
        wedges, texts, autotexts = canvas.axes.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors[:len(labels)], textprops=dict(color="white"))
        canvas.axes.set_title(title, color="white", fontsize=10, pad=10)
        for autotext in autotexts: autotext.set_color('black'); autotext.set_weight('bold')
        canvas.draw()